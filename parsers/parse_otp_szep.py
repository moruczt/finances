from io import BytesIO
import hashlib
import json
import re

from openpyxl import load_workbook
from fastapi import UploadFile
from sqlalchemy import select, insert

import models
from utils import log


UNKNOWN_ACCOUNT_ID = 30


async def parse_otp_szep(file:UploadFile, db, import_id:int, account_id:int) -> dict:
    hash_cols = ["Dátum","Alszámla","Jóváírás","Terhelés","Ellenoldali név","Jogcím"]
    tr_cols = {"date":"Dátum"}

    headers = {}
    data = []
    with BytesIO(await file.read()) as f:
        wb = load_workbook(f, read_only=True, data_only=True, keep_links=False)
        ws = wb[wb.sheetnames[0]]

        for r, row in enumerate(ws.iter_rows(max_col=8, values_only=True)):
            tr = {}
            for c, col in enumerate(row):
                if not r:
                    headers[c] = col
                else:
                    tr[headers[c]] = col
            if r:
                data.append(tr)
    
    
    tr_data = {}
    hashes = []
    fingerprints = []
    for tr in data:
        hash_base = "|".join((str(tr[hc]) for hc in hash_cols))
        hash = hashlib.sha256(hash_base.encode()).hexdigest()
        occurence = hashes.count(hash)
        fingerprint_base = f"{hash_base}|{occurence}"
        fingerprint = hashlib.sha256(fingerprint_base.encode()).hexdigest()
        fingerprints.append(fingerprint)
        tr_data[fingerprint] = {"raw_data":tr,
                                "tr_data":{"date":tr[tr_cols["date"]],
                                           "description":"",
                                           "source":"import",
                                           "is_temporary":False},
                                "entries_data":[{"transaction_id":None,
                                                 "account_id":account_id,
                                                 "raw_import_id":None,
                                                 "amount_huf":(tr["Jóváírás"] or 0) - (tr["Terhelés"] or 0),
                                                 "amount_orig":(tr["Jóváírás"] or 0) - (tr["Terhelés"] or 0)},
                                                {"transaction_id":None,
                                                 "account_id":None,
                                                 "raw_import_id":None,
                                                 "amount_huf":-((tr["Jóváírás"] or 0) - (tr["Terhelés"] or 0)),
                                                 "amount_orig":-((tr["Jóváírás"] or 0) - (tr["Terhelés"] or 0))}]}

    query = select(models.RawImport.row_hash).where(models.RawImport.row_hash.in_(fingerprints))
    existings = set((await db.execute(query)).scalars())

    query = select(models.Rule.target_account_id, models.Rule.conditions).where(models.Rule.account_id==account_id)
    rules = {r["target_account_id"]:json.loads(r["conditions"]) for r in (await db.execute(query)).mappings().all()}
    dates = []
    for row_hash, data in tr_data.items():
        raw_data_json = json.dumps(data["raw_data"])
        if row_hash in existings:
            log(f"HASH FOUND: {row_hash}\n{raw_data_json}")
            continue
        dates.append(data["tr_data"]["date"])

        query = insert(models.RawImport).values(account_id=account_id,
                                                raw_data=raw_data_json,
                                                row_hash=row_hash,
                                                import_id=import_id).returning(models.RawImport.id)
        raw_import_id = (await db.execute(query)).scalar_one()

        query = select(models.AccountConfig.account_id)
        transfer_account_ids = (await db.execute(query)).scalars()

        categorized = False
        match_count = 0
        for target_account_id, conditions in rules.items():
            fail = False
            for col, val in conditions.items():
                if not re.search(val, data["raw_data"][col]):
                    fail = True
                    break
            if not fail:
                match_count += 1
                if categorized:
                    continue
                categorized = True
                if target_account_id in transfer_account_ids:
                    pass
                else:
                    query = insert(models.Transaction).values(data["tr_data"]).returning(models.Transaction.id)
                    tr_id = (await db.execute(query)).scalar_one()

                    [ed.update(transaction_id=tr_id, raw_import_id=raw_import_id) for ed in data["entries_data"]]
                    data["entries_data"][-1]["account_id"] = target_account_id
                    query = insert(models.Entry).values(data["entries_data"])
                    await db.execute(query)


        
        if not categorized:
            data["tr_data"]["is_temporary"] = True
            query = insert(models.Transaction).values(data["tr_data"]).returning(models.Transaction.id)
            tr_id = (await db.execute(query)).scalar_one()

            [ed.update(transaction_id=tr_id, raw_import_id=raw_import_id) for ed in data["entries_data"]]
            data["entries_data"][-1]["account_id"] = UNKNOWN_ACCOUNT_ID
            query = insert(models.Entry).values(data["entries_data"])
            await db.execute(query)

    min_date = min(dates) if dates else None
    max_date = max(dates) if dates else None
    return {"success":True, "row_count":len(fingerprints), "imported_count":len(fingerprints)-len(existings), "min_date":min_date, "max_date":max_date}